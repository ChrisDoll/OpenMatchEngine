# %%
#!/usr/bin/env python3
"""

FM .jsb quick-spec

    Key header

        1 byte inline length (0x01-0x60) or prefix 0x2A/4A/5A/6A ➜ next byte = length.

        then UTF-8 key text.

    Value marker

        0x02 → little-endian INT32 (4 bytes)

        0x03 → little-endian INT64 (8 bytes)

        0x08 → long STRING: u32 len + bytes

        0x80-0x8F → short STRING: low-nibble = len (1-15)

        anything else (e.g. 0xC9, 0x99, 0x5A) = container/control marker.

    Season (one of six in values)

expected_score_data  : ARRAY(11) of {name,str ; negative_multiplier,int ; positive_multiplier,int}
role_data            : ARRAY(var) of blocks
    └─ each block has 'coefficients' ARRAY(52) of {name,str ; value,int}
role_lookup_data     : ARRAY(45) of {index,int ; role,int}
start_value          : INT32
version              : {version_major/minor/release/year : INT32x4}

Offsets (FM 24 season)

expected_score_data : 0x00067725  (start of key text)
role_data           : 0x00067AAE
role_lookup_data    : 0x000760C1
start_value         : 0x000764C5
version             : 0x000764D6

Begin parsing data 19 bytes after the expected_score_data key, skip one container byte (0xC9), then read the 11 triplets.

Typical decode errors - if the tokenizer mis-classifies 0x88 (short-string) as long-string, the stream desynchronises and the next “int” read returns a control byte like <m0x6E>.

season24_decoder.py  —  FM24 .jsb → season-only JSON
----------------------------------------------------
• Hard-seeks to byte offset 0x00067723 (start of season with version_year 24).
• Decodes the single season object:
      expected_score_data  (11 triplets)
      role_data            (coeff-blocks until role_lookup_data)
      role_lookup_data     (45 index/role rows)
      start_value          (int32)
      version { major, minor, release, year }
• Saves output to player_ratings_data_season24.json
• Double-click friendly (console pause at the end).

player_ratings_decoder.py  —  FM24 *.jsb → *.json

player_ratings_data.jsb
└─ values                                        # ARRAY -- 6 season-objects
   ├─ expected_score_data                        # ARRAY[11]
   │     ┌───────────────────────────────────────────────────────────┐
   │     │ Triplet  i  (same for all 11 rows)                       │
   │     │ ──────────────────────────────────────────────────────── │
   │     │   name                   : STRING                        │
   │     │   negative_multiplier    : INT32                         │
   │     │   positive_multiplier    : INT32                         │
   │     └───────────────────────────────────────────────────────────┘
   │            e.g.  "Exceed Significant with Win", 1000, 1200
   │                  "Neutral", 1000, 1000
   │                  …  (11 total rows - identical across seasons)
   │
   ├─ role_data                                   # ARRAY[26 … 38] *
   │     Season-specific performance buckets.
   │
   │     └─ (block k)                             # OBJECT
   │         └─ coefficients                      # ARRAY[52]
   │              • { name: STRING, value: INT32 }
   │
   │              Examples from the first season:
   │              ─ "90min Adjuster"                       →  -650
   │              ─ "Goals"                                →   600
   │              ─ "Good Goal Adder" / "Great Goal Adder" →   100 / 200
   │              ─ …  48 others - **always 52 per block**
   │
   │         *Number of coefficient-blocks per season
   │          (k above) varies:
   │              season 0 … 35 blocks
   │              season 1 … 26
   │              season 2 … 35
   │              season 3 … 35
   │              season 4 … 38
   │              season 5 … 38
   │
   ├─ role_lookup_data                           # ARRAY[45]
   │     { index: INT32,  role: INT32 }          # role is a bit-mask
   │     (identical 0 - 44 mapping in every season)
   │
   ├─ start_value      : INT32                   # e.g. 6700, 6500, …
   │
   └─ version                                     # OBJECT
         ├─ version_major   : INT32   (always 0)
         ├─ version_minor   : INT32   (always 0)
         ├─ version_release : INT32   (6)
         └─ version_year    : INT32   (21)

[
    {1,Goalkeeper},
    {2,Central Defender},
    {4,Full Back},
    {8,Wing Back},
    {16,Defensive Midfielder},
    {32,Central Midfielder},
    {64,Wide Midfielder},
    {128,Winger},
    {512,Attacking Midfielder},
    {1024,Deep Lying Forward},
    {2048,Advanced Forward},
    {4096,Sweeper Keeper},
    {32768,Deep Lying Playmaker},
    {65536,Box to Box Midfielder},
    {131072,Advanced Playmaker},
    {262144,Target Forward},
    {524288,Poacher},
    {1048576,Complete Forward},
    {8388608,,Ball Playing Defender},
    {134217728,Inside Forward},
    {268435456,Ball winning midfielder},
    {536870912,No Nonsense CB},
    {1073741824,Defensive Winger},
    {2147483648,Pressing Forward},
    {4294967296,Trequartista},
    {8589934592,Anchor},
    {68719476736,No Nonsense Full Back},
    {137438953472,Enganche},
    {274877906944,Complete Wing Back},
    {549755813888,Regista},
    {1099511627776,False Nine},
    {2199023255552,Shadow Striker},
    {4398046511104,Wide Target Forward},
    {8796093022208,Wide Playmaker},
    {17592186044416,Inverted Wing Back},
    {35184372088832,Raumdeuter},
    {70368744177664,Roaming Playmaker},
    {140737488355328,Mezzala},
    {281474976710656,Carrilero},
    {1125899906842624,Segundo Volante},
    {2251799813685248,Wide Centre Back},
    {4503599627370496,Inverted full back}
]

"""

import struct
import json
from pathlib import Path
from dataclasses import dataclass, asdict
import sys
from collections import defaultdict
from openpyxl import Workbook  # single lightweight dependency
import importlib.util
import importlib
import subprocess

def ensure(pkg: str):
    """
    Import *pkg*, installing it with pip only if the user agrees.

    1 - install and continue
    2 - exit
    """
    if importlib.util.find_spec(pkg) is None:
        choice = input(
            f"The following script relies on {pkg} to work.\n"
            "Press 1 to install and continue, press 2 to exit: "
        ).strip()
        if choice == "1":
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
            except subprocess.CalledProcessError:
                print(f"Installation of {pkg} failed. Exiting.")
                sys.exit(1)
        else:
            print("Exiting as requested.")
            sys.exit(0)

    globals()[pkg] = importlib.import_module(pkg)

@dataclass
class Hex_address_object:
    """
    Hex address locations from player_ratings_data.jsb

    In each case the first byte of the relevant data structure
    is at the given hex address.
    """

    expected_score_data: str
    role_data: str
    role_lookup_data: str
    start_value: str
    version: str


### HARD CODED ADRESSES OF EACH SEASON IN THE ORIGINAL .jsb FILE ###
fm24_season = Hex_address_object(
    expected_score_data="0x00067725",
    role_data="0x00067AAE",
    role_lookup_data="0x000760C1",
    start_value="0x000764C5",
    version="0x000764D6",
)

fm2302_season = Hex_address_object(
    expected_score_data="0x0005977F",
    role_data="0x00059B08",
    role_lookup_data="0x000672BD",
    start_value="0x000676C1",
    version="0x000676D2",
)
fm2301_season = Hex_address_object(
    expected_score_data="0x0004CD78",
    role_data="0x0004D101",
    role_lookup_data="0x0005932B",
    start_value="0x0005971F",
    version="0x00059730",
)


@dataclass
class Expected_score_object:
    """A single expected_score_data triplet from player_ratings_data.jsb"""

    name: str
    negative_multiplier: int
    positive_multiplier: int


@dataclass
class Role_object:
    """A single role_data coefficient from player_ratings_data.jsb"""

    name: str
    value: int


@dataclass
class Role_lookup_object:
    """A single role_lookup_data entry from player_ratings_data.jsb"""

    index: int
    role: int


@dataclass
class RatingsObject:
    """A single season object from player_ratings_data.jsb"""

    locations: Hex_address_object  # Hex address locations of the fields
    expected_score_data: list[Expected_score_object]
    role_data: list[Role_object]
    role_lookup_data: list[Role_lookup_object]
    start_value: int
    version: dict


# Each expected_score_data entry is a triplet:


#### ─── endianness helpers ───────────────────────────────────────────
def i32(b):
    return struct.unpack_from("<i", b)[0]


def i64(b):
    return struct.unpack_from("<q", b)[0]


def u32(b):
    return struct.unpack_from("<I", b)[0]


def hex_to_int(h):
    return int(h, 16) if isinstance(h, str) else h


def dump_bytes(buf: bytes, pos: int, win: int = 64) -> None:
    """
    Print a hex/ASCII window (±win bytes) centred on *pos*.
    16-byte rows, offsets prefixed, non-printables → '.'.
    """
    lo, hi = max(0, pos - win), min(len(buf), pos + win)
    for offset in range(lo, hi, 16):
        chunk = buf[offset : offset + 16]
        hex_part = " ".join(f"{b:02X}" for b in chunk).ljust(47)
        ascii_part = "".join(chr(b) if 32 <= b <= 126 else "." for b in chunk)
        marker = "◀" if offset <= pos < offset + 16 else " "
        print(f"{marker} 0x{offset:08X}: {hex_part}  {ascii_part}")
    print()


# ─── token reader (length-prefixed keys, no checksum skipping) ───
def tok_stream(buf: bytes, start: int, stop: int):
    """
    Yields (key, typ, val, cursor) until cursor >= stop.
    typ ∈ {"i32","i64","str","ctl"}.
    """
    i, n = start, len(buf) if stop is None else stop
    while i + 2 < n:
        mark = buf[i]

        # key length
        if mark in (0x2A, 0x4A, 0x5A, 0x6A):
            i += 1
            klen = buf[i]
            i += 1
            if not 0 < klen <= 96:
                continue
        elif 0 < mark <= 96:
            klen = mark
            i += 1
        else:
            i += 1
            continue

        key = buf[i : i + klen].decode("utf-8", "replace")
        i += klen
        vmark = buf[i]
        i += 1

        if vmark == 0x02:
            val, typ = i32(buf[i : i + 4]), "i32"
            i += 4
        elif vmark == 0x03:
            val, typ = i64(buf[i : i + 8]), "i64"
            i += 8
        elif vmark == 0x08:
            slen = u32(buf[i : i + 4])
            i += 4
            val = buf[i : i + slen].decode("utf-8", "replace")
            i += slen
            typ = "str"
        elif 0x80 <= vmark <= 0x8F:
            slen = vmark & 0x0F
            val = buf[i : i + slen].decode("utf-8", "replace")
            i += slen
            typ = "str"
        else:
            val, typ = f"<m0x{vmark:02X}>", "ctl"
        yield key, typ, val, i
        if i >= n:
            break




def next_token(ts, wanted_key, wanted_typ, ctx: str):
    try:
        k, t, v, _ = next(ts)
    except StopIteration:
        raise RuntimeError(f"⛔  Unexpected EOF while reading {ctx}")
    if (k, t) != (wanted_key, wanted_typ):
        raise RuntimeError(
            f"⛔  Token mismatch in {ctx}: "
            f"wanted ({wanted_key},{wanted_typ}) - got ({k},{t})"
        )
    return v


def expect(ts, wanted_key, wanted_typ, ctx, buf, *, dump_win=16) -> tuple:
    """
    Fetch next token from *ts* and assert it matches (*wanted_key*, *wanted_typ*).

    wanted_typ can be a single string or a tuple of allowed types.
    On mismatch: prints a contextual hex/ASCII dump of ±dump_win bytes and raises.
    Returns (value, cursor_pos).
    """
    try:
        k, t, v, cur = next(ts)
    except StopIteration:
        raise RuntimeError(f"EOF inside {ctx}")

    if isinstance(wanted_typ, str):
        type_ok = t == wanted_typ
    else:
        type_ok = t in wanted_typ

    if k != wanted_key or not type_ok:
        print(
            f"[{ctx}] TOKEN MISMATCH  - got ({k},{t}) ‡ expected ({wanted_key},{wanted_typ})"
        )
        dump_bytes(buf, cur, dump_win)
        raise RuntimeError(f"Token mismatch in {ctx}")

    return v, cur


def body_start(buf: bytes, anchor: int, key) -> int:
    """
    Given the address of the first byte in
    the ASCII key return the address of the first token that follows the
    key's container marker (usually 0xC9).
    """

    if buf[anchor : anchor + len(key)] != key:
        raise ValueError("Anchor does not point at expected key text")

    cursor = anchor + len(key)  # byte right after the final 'a'
    ctl = buf[cursor]
    if ctl not in (0xC9, 0x99):  # formats seen so far
        raise ValueError(f"Unexpected container marker 0x{ctl:02X}")
    return cursor + 1  # first real token


def parse_expected_score(buf: bytes, start: int, stop: int, verbose: bool):
    """
    Decode the 11 expected_score triplets with extra debugging chatter.
    Layout per row:
        name                 : STRING
        negative_multiplier  : INT32/INT64
        positive_multiplier  : INT32/INT64
    """
    print("Parsing expected_score_data…")
    key = b"expected_score_data"
    body = body_start(buf, start, key)  # existing helper
    if verbose:
        print(f"Parsing expected_score_data… (start @0x{start:08X}, stop @0x{stop:08X})")
        dump_bytes(buf, start, 64)

    ts = tok_stream(buf, body, stop)  # existing generator
    triplets = []

    for row in range(11):
        ctx = f"expected_score row {row:02d}"

        name, n_pos = expect(ts, "name", "str", ctx, buf)
        if name and ord(name[-1]) < 32:
            if verbose:
                print(
                    f"{ctx}: trimming stray 0x{ord(name[-1]):02X} "
                    "and rewinding 1 byte to resynchronise"
                )
                name = name[:-1]  # drop control char
            ts = tok_stream(buf, n_pos - 1, stop)  # back up stream
        if verbose:
            print(f'{ctx}: name@0x{n_pos:08X} = "{name}"')

        neg, neg_pos = expect(ts, "negative_multiplier", ("i32", "i64"), ctx, buf)
        if verbose:
            print(f"{ctx}: neg_multiplier@0x{neg_pos:08X} = {neg}")

        pos, pos_pos = expect(ts, "positive_multiplier", ("i32", "i64"), ctx, buf)
        if verbose:
            print(f"{ctx}: pos_multiplier@0x{pos_pos:08X} = {pos}")

        triplets.append(Expected_score_object(name, neg, pos))

    print("✓ finished expected_score_data\n")
    return triplets


def _read_varint(buf_: bytes, p_: int):
    res, shift = 0, 0
    while True:
        b = buf_[p_]
        res |= (b & 0x7F) << shift
        p_ += 1
        if b < 0x80:  # last byte
            break
        shift += 7
    val = (res >> 1) ^ -(res & 1)  # ZigZag decode
    return val, p_


def parse_role_data(buf: bytes, start: int, stop: int, verbose: bool):
    """
    Robust parser for the *role_data* section.

    • Handles container markers 0xC9/0x99 (object) and 0x09 (array).
    • Decodes value field as: INT32, INT64, 1-byte tiny±63, or ZigZag varint.
    • Cleans label headers: 0x08 long, 0x80-0xFF short, plus special 0x68.
    """

    KEY = b"role_data"
    klen = len(KEY)
    print(f"Parsing role_data… (start @0x{start:08X}, stop @0x{stop:08X})")

    if verbose:
        dump_bytes(buf, start, 64)

    # ── locate body start ───────────────────────────────────────
    if buf[start : start + klen] != KEY:
        raise RuntimeError("role_data key text not found")

    marker = buf[start + klen]
    if marker in (0xC9, 0x99):  # object container
        body = start + klen + 1
    elif marker == 0x09:  # array container
        body = start + klen + 5  # skip u32 length
        if verbose:
            print(f"Array marker 0x09 - body @0x{body:08X}")
    else:
        if verbose:
            dump_bytes(buf, start + klen, 32)
        raise RuntimeError(f"role_data: unknown container 0x{marker:02X}")

    data = buf[body:stop]
    NAME_S = b":\x04name"
    VALUE_S = b"\x05value"

    # ── helpers ─────────────────────────────────────────────────
    def _clean_label(raw: bytes) -> str:
        if not raw:
            return ""
        tag = raw[0]
        if tag == 0x08 and len(raw) >= 5:  # long
            raw = raw[5:]
        elif tag >= 0x80 or tag == 0x68:  # short / special 0x68
            raw = raw[1:]
        else:  # fallback
            while raw and raw[0] < 32:
                raw = raw[1:]
        return raw.decode("utf-8", "replace")

    # ── main scan ───────────────────────────────────────────────
    blocks, coeffs = [], []
    cur, blk_id = 0, 0

    while True:
        npos = data.find(NAME_S, cur)
        if npos == -1:
            break

        # label
        lstart = npos + len(NAME_S)
        vpos = data.find(VALUE_S, lstart)
        if vpos == -1:
            dump_bytes(buf, body + lstart, 64)
            break
        name = _clean_label(data[lstart:vpos])

        # value
        p = vpos + len(VALUE_S)
        tag = data[p]

        if tag == 0x02:  # INT32
            value = struct.unpack_from("<i", data, p + 1)[0]
            nxt, tag_desc = p + 5, "i32"

        elif tag == 0x03:  # INT64
            value = struct.unpack_from("<q", data, p + 1)[0]
            nxt, tag_desc = p + 9, "i64"

        elif 0x80 <= tag <= 0xBF:  # tiny positive
            value = tag - 0x80
            nxt, tag_desc = p + 1, "tiny+"

        elif 0xC0 <= tag <= 0xFF:  # tiny negative
            value = -(tag - 0xC0)
            nxt, tag_desc = p + 1, "tiny−"

        else:  # ZigZag varint
            value, nxt = _read_varint(data, p)
            tag_desc = "var"

        coeffs.append(Role_object(name, value))
        if verbose:
            print(
                f"  block {blk_id:02d}/coeff{len(coeffs) - 1:02d} "
                f"@0x{body + npos:08X}: {name!r:40} = {value:6d}  ({tag_desc})"
            )

        if len(coeffs) == 52:  # seal one block
            blocks.append(coeffs)
            coeffs = []
            blk_id += 1

        cur = nxt
        if body + cur >= stop:
            break

    if coeffs:
        print(f"⚠ trailing fragment with {len(coeffs)} coefficient(s) ignored")

    print(f"✓ finished role_data - {len(blocks)} complete blocks\n")
    return blocks


# ────────────────────────────────────────────────────────────────
# robust, spec-compliant parse_role_lookup
# ────────────────────────────────────────────────────────────────
INDEX_SENT = b"\x05index"  # length-prefixed key (no leading ':')
ROLE_SENT = b"\x04role"

def parse_role_lookup(
    buf: bytes,
    start: int,
    stop: int,
    verbose: bool = False
) -> list[Role_lookup_object]:
    """
    Parse the `role_lookup_data` table in FM24 *.jsb files.

    • Accepts either object (0xC9/0x99) or array (0x09) containers.  
    • Respects the declared row-count when the container is an array;  
      this prevents the parser from drifting into the next section.  
    • Decodes all integer encodings the game uses:
          - 0x02 / 0x03  →  little-endian INT32 / INT64
          - 0x04 / 0x05  →  little-endian UINT32 / UINT64
          - 0x80-0xFF    →  “tiny unsigned” 0-127
          - 0x8? nibble  →  special 1-byte tag (value 0-15)  **← new**
    Returns a list of `Role_lookup_object(index, role)`.
    """
    KEY = b"role_lookup_data"
    if buf[start : start + len(KEY)] != KEY:
        raise RuntimeError("role_lookup_data key not found")

    # ── locate body start & determine declared length ─────────────
    ctag = buf[start + len(KEY)]
    if ctag in (0xC9, 0x99):                      # object container
        declared = None
        body = start + len(KEY) + 1
    elif ctag == 0x09:                            # array container
        declared = struct.unpack_from("<I", buf, start + len(KEY) + 1)[0]
        body = start + len(KEY) + 5
    else:
        raise RuntimeError(f"unknown container tag 0x{ctag:02X}")

    data = buf[body:stop]
    INDEX_SENT = b"\x05index"     # literal “:index”
    ROLE_SENT  = b"\x04role"      # literal “:role”

    # ── helpers ──────────────────────────────────────────────────
    def _tiny_nibble(tag: int) -> int:
        """Decode 1-byte tag-nibble ints (low-nibble == 0x2)."""
        return (tag & 0x7F) >> 4         # 0x8? → 0-15

    def _read_int(d: bytes, p: int) -> tuple[int, int]:
        """Return (value, next_offset) starting at *p*."""
        tag = d[p]
        if tag >= 0x80 and (tag & 0x0F) == 0x02:        # new nibble codec
            return _tiny_nibble(tag), p + 1
        if tag == 0x02:
            return struct.unpack_from("<i", d, p + 1)[0], p + 5
        if tag == 0x03:
            return struct.unpack_from("<q", d, p + 1)[0], p + 9
        if tag == 0x04:
            return struct.unpack_from("<I", d, p + 1)[0], p + 5
        if tag == 0x05:
            return struct.unpack_from("<Q", d, p + 1)[0], p + 9
        if tag >= 0x80:                                # tiny unsigned
            return tag - 0x80, p + 1
        raise ValueError(f"bad int tag 0x{tag:02X} at 0x{p:08X}")

    # ── main scan loop ──────────────────────────────────────────
    rows, cur = [], 0
    while True:
        try:
            idx_off = data.index(INDEX_SENT, cur)
        except ValueError:
            break  # no more “:index”
        index, pos_after_idx = _read_int(data, idx_off + len(INDEX_SENT))

        role_off = data.index(ROLE_SENT, pos_after_idx)
        role, pos_after_role = _read_int(data, role_off + len(ROLE_SENT))

        rows.append(Role_lookup_object(index, role))
        if declared is not None and len(rows) >= declared:
            break  # we’ve read the whole array - stop here
        cur = pos_after_role

    # sanity-check when the container declared a length
    if declared is not None and len(rows) != declared:
        raise RuntimeError(
            f"expected {declared} rows, got {len(rows)} (data truncated or malformed)"
        )

    # optional verbose dump
    if verbose:
        for i, r in enumerate(rows):
            print(f"[{i:02d}]  index={r.index:<3d}  role={r.role}")

    return rows
# ────────────────────────────────────────────────────────────────
# start_value  - single INT32
# ────────────────────────────────────────────────────────────────
def parse_start_value(buf: bytes, anchor: int, verbose: bool = False) -> int:
    """
    Decode the scalar `start_value field situated at *anchor*,
    where *anchor* is the first byte of the ASCII key text.

    Returns the integer value (INT32/INT64, little-endian).

    Raises RuntimeError on any format violation.
    """
    KEY = b"start_value"
    klen = len(KEY)
    ctx = "start_value"
    print("Parsing start_value")
    # 1) Sanity-check that we’re really on the key
    if buf[anchor : anchor + klen] != KEY:
        raise RuntimeError(f"{ctx}: key text not found at 0x{anchor:08X}")

    # 2) Address of the value-marker byte (0x02 = INT32, 0x03 = INT64)
    pos = anchor + klen
    vmark = buf[pos]
    pos += 1  # now *pos* points at the first data byte (if any)

    # 3) Decode
    if vmark == 0x02:  # INT32
        value = struct.unpack_from("<i", buf, pos)[0]
        end = pos + 4
    elif vmark == 0x03:  # INT64
        value = struct.unpack_from("<q", buf, pos)[0]
        end = pos + 8
    else:
        raise RuntimeError(
            f"{ctx}: unexpected value marker 0x{vmark:02X} at 0x{anchor + klen:08X}"
        )

    # 4) Optional verbose diagnostics
    print(f"{ctx}: found {value} (0x{value:08X}) at 0x{pos:08X} - 0x{end - 1:08X}")

    return value


# ────────────────────────────────────────────────────────────────
# version  - 4 keyed INT32 fields
# ────────────────────────────────────────────────────────────────
def parse_version(buf: bytes, anchor: int, verbose: bool = False) -> dict:
    """
    Decode the block

        "version": {
            version_major,
            version_minor,
            version_release,
            version_year
        }

    starting with the byte *anchor* that holds the “v” in “version”.
    Returns a dict with those four integers (all non-negative).

    Tiny-integer rule:
        0x80-0xFF  → unsigned 0-63  (value = tag & 0x3F)
    """
    print("Parsing version data…")
    # ── helper ──────────────────────────────────────────────────
    def _decode_int(buf: bytes, p: int):
        """Return (value, next_offset) for the int starting at *p*."""
        tag = buf[p]

        # tiny unsigned 0-63  (both 0x80-0xBF and 0xC0-0xFF)
        if tag >= 0x80:
            return tag & 0x3F, p + 1

        # 32-bit signed
        if tag == 0x02:
            return struct.unpack_from("<i", buf, p + 1)[0], p + 5

        # 64-bit signed
        if tag == 0x03:
            return struct.unpack_from("<q", buf, p + 1)[0], p + 9

        raise RuntimeError(f"unknown int tag 0x{tag:02X} at 0x{p:08X}")

    # ── sanity check & skip outer “version” key ────────────────
    KEY_VERSION = b"version"
    if buf[anchor : anchor + len(KEY_VERSION)] != KEY_VERSION:
        raise RuntimeError("parse_version: anchor not on 'version' key")

    pos = anchor + len(KEY_VERSION)  # past key text
    if buf[pos] != 0x5A:  # object container marker
        raise RuntimeError("parse_version: expected 0x5A after 'version'")
    pos += 1

    wanted = (
        b"version_major",
        b"version_minor",
        b"version_release",
        b"version_year",
    )
    out = {}

    for k in wanted:
        # ── read key header ─────────────────────────────────────
        klen = buf[pos]
        pos += 1
        if klen != len(k):
            raise RuntimeError(f"unexpected key length {klen} at 0x{pos - 1:08X}")
        if buf[pos : pos + klen] != k:
            got = buf[pos : pos + klen]
            raise RuntimeError(f"got wrong key {got!r} at 0x{pos:08X}")
        pos += klen

        # ── value ───────────────────────────────────────────────
        val, pos = _decode_int(buf, pos)
        out[k.decode()] = val
        if verbose:
            print(f"{k.decode():16} = {val}")

    print("✓ finished version\n")
    return out


def ratingsobject_to_dict(r_obj) -> dict:
    """
    Transform a *RatingsObject* into the exact JSON structure requested.

    • Drops the internal *locations* field.
    • Wraps every 52-coeff list from role_data as
          { "coefficients": [ {name, value}, … ] }
    """
    return {
        "expected_score_data": [asdict(t) for t in r_obj.expected_score_data],
        "role_data": [
            {"coefficients": [asdict(c) for c in coeff_block]}
            for coeff_block in r_obj.role_data
        ],
        "role_lookup_data": [asdict(r) for r in r_obj.role_lookup_data],
        "start_value": r_obj.start_value,
        "version": r_obj.version,
    }


# ─── DECODE ONE SEASON GIVEN ITS OFFSETS ──────────────────────────
def decode_season(buf: bytes, loc: Hex_address_object, verbose=False) -> RatingsObject:
    es_start = hex_to_int(loc.expected_score_data)
    rd_start = hex_to_int(loc.role_data)
    rl_start = hex_to_int(loc.role_lookup_data)
    sv_start = hex_to_int(loc.start_value)
    ver_start = hex_to_int(loc.version)

    print(
        f"Decoding season data at offsets:\n"
        f"  expected_score_data: {es_start:#010x}\n"
        f"  role_data          : {rd_start:#010x}\n"
        f"  role_lookup_data   : {rl_start:#010x}\n"
        f"  start_value        : {sv_start:#010x}\n"
        f"  version            : {ver_start:#010x}"
    )

    if not (
        0 <= es_start < len(buf)
        and 0 <= rd_start < len(buf)
        and 0 <= rl_start < len(buf)
        and 0 <= sv_start < len(buf)
        and 0 <= ver_start < len(buf)
    ):
        raise ValueError("Invalid offsets for the given season data.")

    expected = parse_expected_score(buf, es_start, rd_start, verbose=verbose)
    role = parse_role_data(buf, rd_start, rl_start, verbose=verbose)
    rlookup = parse_role_lookup(buf, rl_start, sv_start, verbose=verbose)
    startval = parse_start_value(buf, sv_start, verbose=verbose)
    version = parse_version(buf, ver_start, verbose=verbose)

    return RatingsObject(loc, expected, role, rlookup, startval, version)


def _pause(input_msg: str = "Press Enter to continue…"):
    if getattr(sys.flags, "interactive", 0):
        return
    if "ipykernel" in sys.modules or "IPython" in sys.modules:
        return
    try:
        input(input_msg)
    except EOFError:
        pass

ROLE_BIT_TO_NAME = {
    1: "Goalkeeper",
    2: "Central Defender",
    4: "Full Back",
    8: "Wing Back",
    16: "Defensive Midfielder",
    32: "Central Midfielder",
    64: "Wide Midfielder",
    128: "Winger",
    512: "Attacking Midfielder",
    1024: "Deep Lying Forward",
    2048: "Advanced Forward",
    4096: "Sweeper Keeper",
    8192: "Libero",
    16384: "Half Back",
    32768: "Deep Lying Playmaker",
    65536: "Box to Box Midfielder",
    131072: "Advanced Playmaker",
    262144: "Target Forward",
    524288: "Poacher",
    1048576: "Complete Forward",
    2097152: "False 9",
    4194304: "Wide Target Forward",
    8388608: "Ball Playing Defender",
    16777216: "Roaming Playmaker",
    33554432: "Mezzala",
    67108864: "Carrilero",
    134217728: "Inside Forward",
    268435456: "Ball Winning Midfielder",
    536870912: "No Nonsense CB",
    1073741824: "Defensive Winger",
    2147483648: "Pressing Forward",
    4294967296: "Trequartista",
    8589934592: "Anchor",
    17179869184: "Wide Playmaker",
    34359738368: "Inverted Wing Back",
    68719476736: "No Nonsense Full Back",
    137438953472: "Enganche",
    274877906944: "Complete Wing Back",
    549755813888: "Regista",
    1099511627776: "False Nine",
    2199023255552: "Shadow Striker",
    4398046511104: "Wide Playmaker",
    8796093022208: "Inverted Winger",
    17592186044416: "Raumdeuter",
    35184372088832: "Half Space Playmaker",
    70368744177664: "Roaming Playmaker",
    140737488355328: "Mezzala",
    281474976710656: "Carrilero",
    562949953421312: "Deep Lying Forward (Support)",
    1125899906842624: "Segundo Volante",
    2251799813685248: "Wide Centre Back",
    4503599627370496: "Inverted Full Back",
}


def _role_header(mask: int) -> str:
    names = [n for bit, n in ROLE_BIT_TO_NAME.items() if mask & bit]
    return "; ".join(names) if names else f"0x{mask:X}"

def _role_numbers_string(mask: int) -> str:
    nums = [str(bit) for bit in ROLE_BIT_TO_NAME.keys() if mask & bit]
    return "; ".join(nums) if nums else f"0x{mask:X}"

def _build_matrix(blocks, lookups):
    idx_to_mask = defaultdict(int)
    for rl in lookups:
        idx_to_mask[rl.index] |= rl.role  # Combine masks per index

    idx_to_header = {i: _role_header(m) for i, m in idx_to_mask.items()}
    idx_to_role_num = {i: _role_numbers_string(m) for i, m in idx_to_mask.items()}
    ordered_idx = sorted(idx_to_header)

    indices = ordered_idx
    role_nums = [idx_to_role_num[i] for i in ordered_idx]
    headers = [idx_to_header[i] for i in ordered_idx]

    matrix = {}
    for idx, block in enumerate(blocks):
        if idx not in idx_to_header:
            continue
        col = idx_to_header[idx]
        for coeff in block:
            matrix.setdefault(coeff.name, {})[col] = coeff.value

    return indices, role_nums, headers, list(matrix), matrix


# ────────────────────────────────────────────────────────────────
# 4.  Main routine — JSON + Excel
# ────────────────────────────────────────────────────────────────
def main():

    if not JSB.exists():
        raise FileNotFoundError(JSB)
    buf = JSB.read_bytes()
    print(f"✓ Read {len(buf):,} bytes")

    season_obj = decode_season(buf, globals()[f"{SEASON_TO_RUN}_season"], verbose=False)

    # 4-A  JSON (unchanged)
    JSON_PATH.write_text(
        json.dumps({"values": [ratingsobject_to_dict(season_obj)]},
                   indent=2, ensure_ascii=False),
        encoding="utf-8"
    )
    print("✓ Saved", JSON_PATH.name)

    # 4-B  Excel workbook (corrected)
    indices, roles, headers, actions, matrix = _build_matrix(
        season_obj.role_data, season_obj.role_lookup_data
    )

    wb = Workbook()              # ✔ create ONE workbook
    ws = wb.active               # ✔ worksheet belongs to the same wb

    # Header rows
    ws.append([""] + indices)    # row-1: raw index
    ws.append([""] + roles)      # row-2: original role number
    ws.append([""] + headers)    # row-3: readable role name(s)
    ws["A1"] = "Index:"
    ws["A2"] = "Role_Num:"
    ws["A3"] = "Role_Name:"
    # Data rows
    for act in actions:
        ws.append([act] + [matrix.get(act, {}).get(h, None) for h in headers])

    wb.save(XLSX_PATH)
    print("✓ Saved", XLSX_PATH.name)

    _pause("Finished - press Enter to exit…")

# ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:

        ensure("openpyxl")

        # ─── Configuration ────────────────────────────────────────────────────
        FILE_NAME = "player_ratings_data.jsb"
        SEASON_TO_RUN = "fm24"  # "fm2302", "fm2301", "fm24"
        ROOT = Path(__file__).parent.parent
        JSB = ROOT / "src" / "clean_simatch" / FILE_NAME
        JSON_PATH = ROOT / "src" / "player_ratings_data.json"
        XLSX_PATH = ROOT / "player_ratings_data.xlsx"

        main()

    except Exception as e:
        print("⛔", e)
        _pause("Press Enter to close…")